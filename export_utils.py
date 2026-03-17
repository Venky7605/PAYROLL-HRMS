from __future__ import annotations

from io import BytesIO
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, GradientFill
from openpyxl.utils import get_column_letter

from company_config import (
    AGE_CATEGORIES, COMPANY_CONFIG, FOOTER_NOTES, PF_OPTIONS, REGIMES, SOURCE_URLS,
)
from tax_engine import (
    CalculatorInputs, RegimeResult, annual_breakup_frame, comparison_table, input_snapshot_frame,
)

# ── Style constants ────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
HDR2_FILL = PatternFill("solid", fgColor="2563EB")
SEC_FILL = PatternFill("solid", fgColor="DBEAFE")
ALT_FILL = PatternFill("solid", fgColor="F0F7FF")
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WHITE_FONT = Font(color="FFFFFF", bold=True)
DARK_FONT = Font(color="1E293B", bold=True)
MONEY_FMT = '₹#,##0.00;[Red](₹#,##0.00);-'
INT_FMT = '₹#,##0;[Red](₹#,##0);-'
PCT_FMT = '0.00"%"'


def _title_row(ws, row: int, text: str, colspan: int = 8, fill=HDR_FILL) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)


def _hdr_cell(ws, row: int, col: int, value: str, fill=HDR2_FILL) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = WHITE_FONT
    c.fill = fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN


def _data_cell(ws, row: int, col: int, value, fmt: str | None = None, fill=None, bold=False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN
    c.alignment = Alignment(
        horizontal="right" if isinstance(value, (int, float)) else "left",
        vertical="center",
    )
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if bold:
        c.font = Font(bold=True, color="1E293B")


def _auto_fit(ws, min_col: int = 1, max_col: int | None = None) -> None:
    if max_col is None:
        max_col = ws.max_column
    for idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(idx)
        max_len = max(
            (len(str(c.value)) for c in ws[col_letter] if c.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 42)


def _write_kv_block(ws, start_row: int, pairs, fill_key=SEC_FILL, fill_val=None) -> int:
    for label, value in pairs:
        ws.cell(row=start_row, column=1, value=label).fill = fill_key
        ws.cell(row=start_row, column=1).font = DARK_FONT
        ws.cell(row=start_row, column=1).border = THIN
        c = ws.cell(row=start_row, column=2, value=value)
        c.border = THIN
        if fill_val:
            c.fill = fill_val
        start_row += 1
    return start_row


def build_export_workbook(
    inputs: CalculatorInputs,
    results: Dict[str, RegimeResult],
    selected_regime_key: str,
    employee_name: str = "",
    employee_id: str = "",
    department: str = "",
    designation: str = "",
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sel = results[selected_regime_key]
    alt_key = "new" if selected_regime_key == "old" else "old"
    alt = results[alt_key]
    comparison_df = comparison_table(results)

    # ── Sheet 1: Summary & Comparison ────────────────────────────────────────
    ws = wb.create_sheet("📊 Summary")
    _title_row(ws, 1, f"{COMPANY_CONFIG.app_title}  —  {inputs.financial_year}", colspan=10)

    row = 3
    emp_pairs = [
        ("Employee Name", employee_name or "—"),
        ("Employee ID", employee_id or "—"),
        ("Department", department or "—"),
        ("Designation", designation or "—"),
        ("Financial Year", inputs.financial_year),
        ("Joining Date", inputs.join_date.isoformat()),
        ("PF Option", PF_OPTIONS[inputs.pf_option]),
        ("Age Category", AGE_CATEGORIES[inputs.age_category]),
        ("Selected Regime", REGIMES[selected_regime_key]),
        ("Recommended Regime", REGIMES[alt_key] if alt.annual_tax < sel.annual_tax else REGIMES[selected_regime_key]),
    ]
    row = _write_kv_block(ws, row, emp_pairs)

    row += 1
    _title_row(ws, row, "REGIME COMPARISON", colspan=len(comparison_df.columns) + 1, fill=HDR2_FILL)
    row += 1

    for col_idx, col_name in enumerate(comparison_df.columns, start=1):
        _hdr_cell(ws, row, col_idx, col_name)
    row += 1

    for r_idx, record in enumerate(comparison_df.itertuples(index=False, name=None)):
        f = GREEN_FILL if r_idx == 0 and sel.regime_key == "old" else (GREEN_FILL if r_idx == 1 and sel.regime_key == "new" else ALT_FILL)
        for col_idx, val in enumerate(record, start=1):
            fmt = INT_FMT if isinstance(val, (int, float)) else None
            _data_cell(ws, row, col_idx, val, fmt=fmt, fill=f)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Key Notes").font = Font(bold=True, size=11)
    row += 1
    for note in FOOTER_NOTES:
        ws.cell(row=row, column=1, value=f"• {note}").font = Font(italic=True, color="475569")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1

    _auto_fit(ws, max_col=10)

    # ── Sheet 2: Inputs Snapshot ──────────────────────────────────────────────
    ws = wb.create_sheet("📋 Inputs")
    _title_row(ws, 1, "Employee Inputs Snapshot", colspan=3)
    input_df = input_snapshot_frame(inputs)
    _hdr_cell(ws, 3, 1, "Input Field")
    _hdr_cell(ws, 3, 2, "Value")
    for row_idx, (label, value) in enumerate(input_df.itertuples(index=False, name=None), start=4):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        _data_cell(ws, row_idx, 1, label, fill=fill, bold=True)
        _data_cell(ws, row_idx, 2, value, fill=fill)
    _auto_fit(ws, max_col=3)

    # ── Sheet 3 & 4: Annual Tax Working ──────────────────────────────────────
    for sheet_label, result in [
        (f"🧮 {result.regime_label[:3]} Tax Working", sel),
        (f"🧮 Alt Tax Working", alt),
    ]:
        result = sel if "Sel" in sheet_label or sheet_label.startswith("🧮 Old") or sheet_label.startswith("🧮 New") else alt
        pass  # redefine below

    for sheet_label, result in [("🧮 Selected Regime", sel), ("🧮 Alternative Regime", alt)]:
        ws = wb.create_sheet(sheet_label)
        _title_row(ws, 1, f"{result.regime_label}  —  Annual Tax Working  —  {result.financial_year}", colspan=6)
        brk = annual_breakup_frame(result)
        _hdr_cell(ws, 3, 1, "Particulars", fill=HDR_FILL)
        _hdr_cell(ws, 3, 2, "Amount (₹)", fill=HDR_FILL)
        for r_idx, (label, amt) in enumerate(brk.itertuples(index=False, name=None), start=4):
            is_total = "=" in label or label.startswith("Net")
            fill = AMBER_FILL if is_total else (ALT_FILL if r_idx % 2 == 0 else None)
            _data_cell(ws, r_idx, 1, label, fill=fill, bold=is_total)
            _data_cell(ws, r_idx, 2, amt, fmt=MONEY_FMT, fill=fill, bold=is_total)

        # Exemptions block
        col = 4
        ws.cell(row=3, column=col, value="Salary Exemptions").font = DARK_FONT
        ws.cell(row=3, column=col).fill = SEC_FILL
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        for r_idx, (k, v) in enumerate(result.salary_exemptions.items(), start=4):
            _data_cell(ws, r_idx, col, k, fill=ALT_FILL)
            _data_cell(ws, r_idx, col + 1, v, fmt=MONEY_FMT, fill=ALT_FILL)

        ws.cell(row=10, column=col, value="Chapter VI-A Deductions").font = DARK_FONT
        ws.cell(row=10, column=col).fill = SEC_FILL
        ws.merge_cells(start_row=10, start_column=col, end_row=10, end_column=col + 1)
        for r_idx, (k, v) in enumerate(result.chapter_via_deductions.items(), start=11):
            fill = GREEN_FILL if v > 0 else None
            _data_cell(ws, r_idx, col, k, fill=fill)
            _data_cell(ws, r_idx, col + 1, v, fmt=MONEY_FMT, fill=fill)

        _auto_fit(ws, max_col=col + 1)

    # ── Sheet 5 & 6: Monthly Details ─────────────────────────────────────────
    for sheet_label, result in [("📅 Selected Monthly", sel), ("📅 Alt Monthly", alt)]:
        ws = wb.create_sheet(sheet_label)
        _title_row(ws, 1, f"{result.regime_label}  —  Monthly Salary Details", colspan=20)
        df = result.monthly_df.copy()
        df["Month Start"] = df["Month Start"].astype(str)
        cols = list(df.columns)
        for col_idx, col_name in enumerate(cols, start=1):
            _hdr_cell(ws, 3, col_idx, col_name)
        for r_idx, row_data in enumerate(df.itertuples(index=False, name=None), start=4):
            fill = ALT_FILL if r_idx % 2 == 0 else None
            for col_idx, val in enumerate(row_data, start=1):
                fmt = MONEY_FMT if isinstance(val, float) and col_idx > 3 else None
                _data_cell(ws, r_idx, col_idx, val, fmt=fmt, fill=fill)
        _auto_fit(ws, max_col=len(cols))

    # ── Sheet 7: Sources ──────────────────────────────────────────────────────
    ws = wb.create_sheet("📎 Sources")
    _title_row(ws, 1, "Official Source Links", colspan=3)
    _hdr_cell(ws, 3, 1, "Reference")
    _hdr_cell(ws, 3, 2, "URL")
    for r_idx, (key, url) in enumerate(SOURCE_URLS.items(), start=4):
        _data_cell(ws, r_idx, 1, key.replace("_", " ").title(), fill=ALT_FILL, bold=True)
        c = ws.cell(row=r_idx, column=2, value=url)
        c.font = Font(color="2563EB", underline="single")
        c.border = THIN
    _auto_fit(ws, max_col=2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
